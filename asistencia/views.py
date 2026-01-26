from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout as django_logout
from django.db.models import Sum
from datetime import timedelta
import json
import openpyxl

# --- IMPORTANTE: Herramientas de diseño para Excel ---
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Importamos tus modelos
from .models import Alumno, Asistencia, Pago

# ========================================================
# 1. SEMÁFORO INTELIGENTE
# ========================================================
@login_required
def smart_login_redirect(request):
    if request.user.is_superuser:
        return redirect('/admin/')
    else:
        return redirect('dashboard')

# ========================================================
# 2. LOGOUT
# ========================================================
def logout(request):
    django_logout(request)
    return redirect('login_asistencia') 

# ========================================================
# 3. DASHBOARD
# ========================================================
@staff_member_required
def dashboard(request):
    total_alumnos = Alumno.objects.count()
    hoy = timezone.now().date()
    asistencias_hoy = Asistencia.objects.filter(fecha__date=hoy).count()
    
    mes_actual = hoy.month
    ingresos_mes = Pago.objects.filter(fecha_pago__month=mes_actual).aggregate(Sum('monto'))['monto__sum']
    if ingresos_mes is None:
        ingresos_mes = 0

    labels = []
    data = []
    for i in range(6, -1, -1):
        fecha_analisis = hoy - timedelta(days=i)
        cnt = Asistencia.objects.filter(fecha__date=fecha_analisis).count()
        labels.append(fecha_analisis.strftime("%d/%m"))
        data.append(cnt)

    context = {
        'total_alumnos': total_alumnos,
        'asistencias_hoy': asistencias_hoy,
        'ingresos_mes': ingresos_mes,
        'chart_labels': json.dumps(labels),
        'chart_data': json.dumps(data),
    }
    return render(request, 'admin/dashboard.html', context)

# ========================================================
# 4. REGISTRO DE ASISTENCIA
# ========================================================
@login_required(login_url='login_asistencia')
def registro_asistencia(request):
    context = {}
    if request.method == 'POST':
        dni_ingresado = request.POST.get('dni')
        try:
            alumno_encontrado = Alumno.objects.get(dni=dni_ingresado)
            if alumno_encontrado.esta_al_dia():
                Asistencia.objects.create(alumno=alumno_encontrado)
                context = {
                    'alumno': alumno_encontrado,
                    'estado': 'exito',
                    'mensaje': f"¡Bienvenido, {alumno_encontrado.nombre}!",
                    'submensaje': f"Vencimiento: {alumno_encontrado.fecha_vencimiento}"
                }
            else:
                context = {
                    'alumno': alumno_encontrado,
                    'estado': 'error',
                    'mensaje': "⛔ ACCESO DENEGADO",
                    'submensaje': f"Membresía vencida el {alumno_encontrado.fecha_vencimiento}"
                }
        except Alumno.DoesNotExist:
            messages.error(request, "❌ DNI no encontrado.")
    return render(request, 'registro.html', context)

# ========================================================
# 5. EXPORTAR EXCEL (ESTILO DASHBOARD FINANCIERO)
# ========================================================
@staff_member_required
def exportar_asistencias_excel(request):
    # 1. Configurar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Asistencias_MCombat.xlsx'

    # 2. Crear libro
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Asistencias"
    
    # --- PALETA DE COLORES ---
    COLOR_NAVY_HEADER = '0F172A'   # Azul Marino muy oscuro (Encabezado Principal)
    COLOR_SUB_HEADER  = '1E293B'   # Gris Azulado (Encabezados de Tabla)
    COLOR_TEXTO_BLANCO= 'FFFFFF'
    COLOR_FILA_PAR    = 'F1F5F9'   # Gris muy clarito (casi blanco)
    COLOR_FILA_IMPAR  = 'FFFFFF'   # Blanco puro
    COLOR_VERDE_EXITO = '22C55E'   # Verde estilo Excel financiero (para resaltar)
    COLOR_BORDE       = 'CBD5E1'   # Gris suave para las líneas

    # --- ESTILOS ---
    # Fuente Título Principal
    fuente_titulo = Font(name='Calibri', size=18, bold=True, color=COLOR_TEXTO_BLANCO)
    relleno_titulo = PatternFill(start_color=COLOR_NAVY_HEADER, end_color=COLOR_NAVY_HEADER, fill_type='solid')

    # Fuente Encabezados de Tabla
    fuente_th = Font(name='Calibri', size=12, bold=True, color=COLOR_TEXTO_BLANCO)
    relleno_th = PatternFill(start_color=COLOR_SUB_HEADER, end_color=COLOR_SUB_HEADER, fill_type='solid')

    # Alineación y Bordes
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left', vertical='center')
    borde  = Border(
        left=Side(style='thin', color=COLOR_BORDE), 
        right=Side(style='thin', color=COLOR_BORDE), 
        top=Side(style='thin', color=COLOR_BORDE), 
        bottom=Side(style='thin', color=COLOR_BORDE)
    )

    # --- ESTRUCTURA DEL REPORTE ---

    # 1. TÍTULO TIPO "BANNER" (Filas 1-2)
    ws.merge_cells('A1:D2')
    celda_titulo = ws['A1']
    celda_titulo.value = "📊 REPORTE DE ASISTENCIAS MCOMBAT"
    celda_titulo.font = fuente_titulo
    celda_titulo.fill = relleno_titulo
    celda_titulo.alignment = center

    # 2. ENCABEZADOS DE COLUMNA (Fila 4 - Dejamos la 3 vacía como separador)
    headers = ['FECHA', 'ALUMNO / LUCHADOR', 'HORA DE LLEGADA', 'DÍA SEMANA']
    fila_header = 4
    
    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=fila_header, column=col_num)
        cell.value = title
        cell.font = fuente_th
        cell.fill = relleno_th
        cell.alignment = center
        cell.border = borde

    # 3. DATOS (Desde Fila 5)
    rows = Asistencia.objects.all().order_by('-id')
    fila_inicial = 5

    for i, row in enumerate(rows):
        num_fila = fila_inicial + i
        
        # Procesar hora
        try:
            hora_str = row.fecha.strftime("%H:%M")
        except:
            hora_str = "-"

        valores = [
            row.fecha.strftime("%Y-%m-%d"), # A: Fecha
            str(row.alumno),                # B: Nombre
            hora_str,                       # C: Hora
            row.fecha.strftime("%A")        # D: Día
        ]

        # Alternar color de fondo (Efecto Cebra)
        if i % 2 == 0:
            color_fondo = COLOR_FILA_PAR
        else:
            color_fondo = COLOR_FILA_IMPAR
        
        relleno_fila = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type='solid')
        fuente_normal = Font(name='Calibri', size=11, color='334155') # Gris oscuro para texto

        for col_num, valor in enumerate(valores, 1):
            cell = ws.cell(row=num_fila, column=col_num)
            cell.value = valor
            cell.fill = relleno_fila
            cell.font = fuente_normal
            cell.border = borde
            
            # Alineación específica
            if col_num == 2: # Nombre del alumno (Columna B)
                cell.alignment = left
                cell.font = Font(name='Calibri', size=11, bold=True, color='334155') # Nombre en negrita
            else:
                cell.alignment = center

            # DETALLE VISUAL: Si es Domingo (Sunday), pintar el texto de ROJO
            if col_num == 4 and valor == 'Sunday':
                 cell.font = Font(name='Calibri', size=11, color='DC2626', bold=True)


    # 4. AUTO-AJUSTE DE ANCHO DE COLUMNAS
    anchos = [15, 45, 20, 20] # Anchos fijos para que se vea ordenado
    letras = ['A', 'B', 'C', 'D']
    
    for i, ancho in enumerate(anchos):
        ws.column_dimensions[letras[i]].width = ancho

    # 5. Guardar
    wb.save(response)
    return response

# Exportar lista de alumnos (Opcional)
@staff_member_required
def exportar_alumnos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alumnos MCombat'
    headers = ['ID', 'Nombre', 'Apellido', 'DNI', 'Fecha Registro']
    ws.append(headers)
    alumnos = Alumno.objects.all()
    for alumno in alumnos:
        fecha_reg = str(alumno.fecha_registro.date()) if alumno.fecha_registro else '-'
        ws.append([alumno.id, alumno.nombre, alumno.apellido, alumno.dni, fecha_reg])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Listado_Alumnos.xlsx"'
    wb.save(response)
    return response

# ========================================================
# 6. CONSULTA PÚBLICA PARA ALUMNOS (SIN LOGIN)
# ========================================================
def consulta_publica(request):
    context = {}
    
    if request.method == 'POST':
        dni_ingresado = request.POST.get('dni')
        
        try:
            alumno = Alumno.objects.get(dni=dni_ingresado)
            
            # Definimos el estado y los colores
            if alumno.esta_al_dia():
                estado_texto = "¡ESTÁS ACTIVO!"
                clase_color = "success" # Verde
                mensaje_extra = "Nos vemos en el entrenamiento. 🥊"
                icono = "✅"
            else:
                estado_texto = "MEMBRESÍA VENCIDA"
                clase_color = "danger"  # Rojo
                mensaje_extra = "Por favor, acércate a recepción para renovar."
                icono = "⛔"
            
            context = {
                'resultado': True,
                'alumno': alumno,
                'estado_texto': estado_texto,
                'clase_color': clase_color,
                'mensaje_extra': mensaje_extra,
                'icono': icono
            }
            
        except Alumno.DoesNotExist:
            context = {
                'error': True,
                'mensaje': "❌ No encontramos ese DNI en el sistema."
            }
            
    return render(request, 'consulta_publica.html', context)